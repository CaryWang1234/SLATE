"""技能：Excel / CSV 办公表格处理（基于 openpyxl + 标准库 csv）。

支持动作：
- create: 生成 .xlsx 表格（headers + rows，或 data 文本按 CSV 格式解析，首行表头）
- read: 读取 .xlsx/.csv 内容（返回表头、数据预览与行列统计）
- convert: csv ↔ xlsx 互转

适用于办公数据整理、报表生成、表格数据查看与格式转换等场景。
输出落盘到数据目录 outputs/，返回 file_path。
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

DATA_DIR = Path(os.environ.get("SLATE_DATA_DIR", Path(__file__).resolve().parent.parent.parent / "data"))

MAX_PREVIEW_ROWS = 50
MAX_CELL_LENGTH = 200


def _outputs_dir() -> Path:
    out = DATA_DIR / "outputs"
    out.mkdir(parents=True, exist_ok=True)
    return out


def _safe_name(name: str) -> str:
    """文件名消毒：去除路径分隔符与非法字符。"""
    return re.sub(r'[\\/:*?"<>|\s]+', "_", name.strip())[:40] or "table"


def _parse_matrix(value: Any) -> list[list[Any]] | None:
    """解析二维数组参数（JSON 字符串或原生 list）。"""
    if value is None or value == "":
        return None
    if isinstance(value, list):
        return [row if isinstance(row, list) else [row] for row in value]
    try:
        data = json.loads(value)
        if isinstance(data, list):
            return [row if isinstance(row, list) else [row] for row in data]
    except (json.JSONDecodeError, TypeError):
        pass
    return None


def _parse_text_rows(text: str) -> list[list[str]]:
    """解析 CSV/TSV 风格文本为二维数组（自动识别分隔符）。"""
    sample = text[:4096]
    delimiter = "\t" if "\t" in sample else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if any(str(c).strip() for c in row)]
    return rows


def _create(title: str, sheet: str, headers: Any, rows: Any, data: str, file_name: str) -> dict[str, Any]:
    """生成 .xlsx 文件。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return {"error": "openpyxl 未安装。请执行: pip install openpyxl"}

    header_list: list[Any] = []
    body: list[list[Any]] | None = None

    if data and str(data).strip():
        body = _parse_text_rows(str(data))
        if not body:
            return {"error": "data 解析为空，请检查格式（每行一条记录，逗号或制表符分隔）"}
        if not headers:
            header_list = body[0]
            body = body[1:]
    else:
        if isinstance(headers, list):
            header_list = headers
        elif headers:
            header_list = [h.strip() for h in str(headers).split(",") if h.strip()]
        body = _parse_matrix(rows)
        if not header_list and not body:
            return {"error": "create 需要 data 文本，或 headers + rows 参数"}

    body = body or []

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet or "").strip()[:31] or "Sheet1"

    if header_list:
        ws.append(header_list)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4A6FA5")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
    
    def _to_cell_value(v):
        """保留数字/布尔等原始类型，仅截断超长字符串。"""
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return v
        s = str(v)[:32000]
        # 尝试还原数字类型
        try:
            if "." in s:
                return float(s)
            return int(s)
        except (ValueError, TypeError):
            return s
    
    for row in body:
        ws.append([_to_cell_value(c) for c in row])

    # 简单列宽自适应（按表头/前若干行估算）
    for idx, col in enumerate(ws.iter_cols(max_row=min(ws.max_row, 20)), start=1):
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 50)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = file_name.strip() if file_name else f"{_safe_name(title or 'table')}_{stamp}.xlsx"
    if not fname.lower().endswith(".xlsx"):
        fname += ".xlsx"
    out_path = _outputs_dir() / fname
    wb.save(str(out_path))
    return {
        "message": "ok",
        "file_path": str(out_path),
        "rows": len(body),
        "columns": len(header_list) or (len(body[0]) if body else 0),
    }


def _read(file_path: str, sheet: str, limit: int) -> dict[str, Any]:
    """读取 .xlsx 或 .csv，返回表头 + 数据预览 + 统计。"""
    p = Path(os.path.expanduser(file_path or ""))
    if not p.is_file():
        return {"error": f"文件不存在: {file_path}"}
    ext = p.suffix.lower()
    cap = min(max(int(limit or MAX_PREVIEW_ROWS), 1), 500)

    if ext == ".csv":
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
        delimiter = "\t" if "\t" in sample else ","
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f, delimiter=delimiter))
        sheet_name = p.stem
    elif ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"error": "openpyxl 未安装。请执行: pip install openpyxl"}
        wb = load_workbook(str(p), read_only=True, data_only=True)
        sheet_name = (sheet or "").strip()
        if sheet_name and sheet_name not in wb.sheetnames:
            err_msg = f"工作表不存在: {sheet_name}，可用: {wb.sheetnames}"
            wb.close()
            return {"error": err_msg}
        ws = wb[sheet_name] if sheet_name else wb.active
        sheet_name = ws.title
        rows = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
        sheets = wb.sheetnames
        wb.close()
    else:
        return {"error": f"不支持的文件类型: {ext}（支持 .xlsx/.csv）"}

    rows = [[str(c)[:MAX_CELL_LENGTH] for c in row] for row in rows]
    headers = rows[0] if rows else []
    body = rows[1:]
    return {
        "status": "ok",
        "file": p.name,
        "sheet": sheet_name,
        "sheets": sheets if ext != ".csv" else None,
        "headers": headers,
        "row_count": len(body),
        "column_count": len(headers) or (len(body[0]) if body else 0),
        "preview_rows": body[:cap],
        "truncated": len(body) > cap,
    }


def _convert(file_path: str, out: str) -> dict[str, Any]:
    """csv → xlsx 或 xlsx → csv。"""
    p = Path(os.path.expanduser(file_path or ""))
    if not p.is_file():
        return {"error": f"文件不存在: {file_path}"}
    ext = p.suffix.lower()

    if ext == ".csv":
        # 直接全量读取（不受预览上限约束）
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
        delimiter = "\t" if "\t" in sample else ","
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f, delimiter=delimiter))
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"error": "openpyxl 未安装。请执行: pip install openpyxl"}
        wb = Workbook()
        ws = wb.active
        ws.title = p.stem[:31] or "Sheet1"
        for row in rows:
            ws.append([_to_cell_value(c) for c in row])
        out_path = Path(out).expanduser() if (out or "").strip() else p.with_suffix(".xlsx")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out_path))
        return {"message": "ok", "file_path": str(out_path), "rows": len(rows)}

    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"error": "openpyxl 未安装。请执行: pip install openpyxl"}
        wb = load_workbook(str(p), read_only=True, data_only=True)
        ws = wb.active
        out_path = Path(out).expanduser() if (out or "").strip() else p.with_suffix(".csv")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        count = 0
        with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if c is None else c for c in row])
                count += 1
        wb.close()
        return {"message": "ok", "file_path": str(out_path), "rows": count}

    return {"error": f"不支持的文件类型: {ext}（支持 .xlsx/.csv 互转）"}


def execute(
    action: str = "",
    file_path: str = "",
    title: str = "",
    sheet: str = "",
    headers: Any = None,
    rows: Any = None,
    data: str = "",
    limit: int = MAX_PREVIEW_ROWS,
    out: str = "",
    file_name: str = "",
    **_kw: Any,
) -> dict[str, Any]:
    """Excel / CSV 办公表格工具。

    Args:
        action: 操作类型 - create/read/convert
        file_path: 源文件路径（read/convert 必填，支持 .xlsx/.csv）
        title: 表格标题（create 时用作文件名，默认 table）
        sheet: 工作表名（create 时命名；read 时指定读取哪个表）
        headers: 表头（create 时使用，JSON 数组或逗号分隔文本）
        rows: 数据行（create 时使用，JSON 二维数组）
        data: CSV 格式文本数据（create 时可用，每行一条记录，逗号/制表符分隔，首行表头）
        limit: 读取预览行数上限（read 时使用，默认 50，最大 500）
        out: 输出路径（convert 时可选，缺省为源同目录同名换扩展名）
        file_name: 输出文件名（create 时可选，缺省为 标题_时间戳.xlsx）

    Returns:
        dict: 操作结果。
    """
    if not action:
        return {"error": "action 不能为空，可选: create/read/convert"}

    try:
        if action == "create":
            return _create(title, sheet, headers, rows, data, file_name)
        if action == "read":
            if not file_path:
                return {"error": "read 操作需要提供 file_path 参数"}
            return _read(file_path, sheet, limit)
        if action == "convert":
            if not file_path:
                return {"error": "convert 操作需要提供 file_path 参数"}
            return _convert(file_path, out)
        return {"error": f"未知操作: {action}，可选: create/read/convert"}
    except Exception as e:
        return {"error": f"表格处理失败 ({action}): {e}"}
